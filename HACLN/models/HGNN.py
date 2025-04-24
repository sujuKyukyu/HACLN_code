from torch import nn
from models import HGNN_conv
import torch.nn.functional as F
import pandas as pd
import openpyxl as ox
from sklearn.preprocessing import normalize
import torch
# HGNN方法改进
# class HGNN(nn.Module):
#     def __init__(self, in_ch, n_class, n_hid, dropout=0.5):
#         super(HGNN, self).__init__()
#         self.dropout = dropout
#         self.hgc_half = HGNN_conv(30, 30)
#         self.hgc1 = HGNN_conv(60, n_hid)
#         self.hgc2 = HGNN_conv(n_hid, n_class)
#
#
#
#     def forward(self, x, G):
#         wb = ox.Workbook()
#         ws = wb.active
#         ws.title = 'test'
#
#         # print(x)
#         # print(x[:,0:30])
#         # print(x[:,0:30].shape)
#         # print(x[:, 30:60])
#         # print(x[:,30:60].shape)
#         x1=x[:,0:30]
#         x2=x[:,30:60]
#         x1=F.dropout(F.relu(self.hgc_half(x1, G)))
#         x2=F.dropout(F.relu(self.hgc_half(x2, G)))
#         x3=F.dropout(F.relu(self.hgc1(x, G)))
#         # print(x3.shape)
#         # print("type(x1)",type(x1))
#         # print("type(x3)",type(x3))
#         # print(x3)
#         # print(((x1*x2)+x3).shape)
#         x = x1*x2+x3
#         # print(x)
#         # print(x.shape)
#         x_30 = x
#         x = self.hgc2(x, G)
#         return x,x_30

# class HGNN(nn.Module):
#     def __init__(self, in_ch, n_class, n_hid, dropout=0.5):
#         super(HGNN, self).__init__()
#         self.dropout = dropout
#         self.hgc1 = HGNN_conv(in_ch, n_hid)
#         self.hgc2 = HGNN_conv(n_hid, n_class)
#
#     def forward(self, x, G):
#         x = F.relu(self.hgc1(x, G))
#         x = F.dropout(x, self.dropout)
#         x_60 = x
#         x = self.hgc2(x, G)
#         print(x_60)
#         return x,x_60

# HGNN方法改进
# class HGNN(nn.Module):
#     def __init__(self, in_ch, n_class, n_hid, dropout=0.5):
#         super(HGNN, self).__init__()
#         self.dropout = dropout
#         self.hgc_half = HGNN_conv(30, 30)
#         self.hgc1 = HGNN_conv(60, n_hid)
#         self.hgc2 = HGNN_conv(n_hid, n_class)
#         # self.fc1 = torch.nn.Linear(n_hid, n_class)
#
#
#
#     def forward(self, x, G):
#         wb = ox.Workbook()
#         ws = wb.active
#         ws.title = 'test'
#
#         # print(x)
#         # print(x[:,0:30])
#         # print(x[:,0:30].shape)
#         # print(x[:, 30:60])
#         # print(x[:,30:60].shape)
#         x1=x[:,0:30]
#         x2=x[:,30:60]
#         x1=F.dropout(F.relu(self.hgc_half(x1, G)))
#         x2=F.dropout(F.relu(self.hgc_half(x2, G)))
#         x3=F.dropout(F.relu(self.hgc1(x, G)))
#         # print(x3.shape)
#         # print("type(x1)",type(x1))
#         # print("type(x3)",type(x3))
#         # print(x3)
#         # print(((x1*x2)+x3).shape)
#         x = x1*x2+x3
#         # print(x)
#         # print(x.shape)
#         x_30 = x
#         x = self.hgc2(x, G)
#         # x_fc1 = self.fc1(x)
#         # x_fc1 = torch.relu(x_fc1)
#         # x_fc1 = F.dropout(x_fc1, 0.6)
#
#         return x,x_30

#HGNN方法改进+对比损失
class HGNN(nn.Module):
    def __init__(self, in_ch, n_class, n_hid, dropout=0.5):
        super(HGNN, self).__init__()
        self.dropout = dropout
        self.hgc_half = HGNN_conv(30, 30)
        self.hgc1 = HGNN_conv(60, n_hid)
        self.hgc2 = HGNN_conv(n_hid, n_class)
        # self.fc1 = torch.nn.Linear(n_hid, n_class)



    def forward(self, x, G):
        wb = ox.Workbook()
        ws = wb.active
        ws.title = 'test'

        # print(x)
        # print(x[:,0:30])
        # print(x[:,0:30].shape)
        # print(x[:, 30:60])
        # print(x[:,30:60].shape)
        x1=x[:,0:30]
        x2=x[:,30:60]
        x1=F.dropout(F.relu(self.hgc_half(x1, G)))
        x2=F.dropout(F.relu(self.hgc_half(x2, G)))
        x3=F.dropout(F.relu(self.hgc1(x, G)))
        # print(x3.shape)
        # print("type(x1)",type(x1))
        # print("type(x3)",type(x3))
        # print(x3)
        # print(((x1*x2)+x3).shape)
        x = x1*x2+x3
        # print(x)
        # print(x.shape)
        x_30 = x
        x = self.hgc2(x, G)
        # x_fc1 = self.fc1(x)
        # x_fc1 = torch.relu(x_fc1)
        # x_fc1 = F.dropout(x_fc1, 0.6)

        return x,x_30